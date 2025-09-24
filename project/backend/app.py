"""
Flask backend for Market Basket Analysis.
Supports market basket analysis using the mlxtend library.
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
import json
import io
import os
from datetime import datetime
from urllib.parse import unquote
import tempfile
from werkzeug.utils import secure_filename, safe_join
import logging
import traceback
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import data processor
from data_processors.basket_analyzer import BasketAnalyzer

# Create Flask app
app = Flask(__name__)
CORS(app)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure application settings
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB
UPLOAD_FOLDER_NAME = 'uploads'
UPLOAD_FOLDER = os.path.join(app.root_path, UPLOAD_FOLDER_NAME)
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

# Ensure the uploads directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def read_file_safely(filepath):
    """Read a file safely using multiple encodings."""
    try:
        if filepath.lower().endswith('.csv'):
            # Try multiple encodings for CSV files
            encodings = ['utf-8', 'utf-8-sig', 'cp1252', 'iso-8859-1', 'tis-620', 'windows-1252']
            for encoding in encodings:
                try:
                    df = pd.read_csv(filepath, encoding=encoding)
                    print(f"[OK] CSV read successfully with encoding: {encoding}")
                    return df
                except UnicodeDecodeError:
                    continue
            
            # Fallback to chardet if the encodings fail
            try:
                import chardet
                with open(filepath, 'rb') as f:
                    raw_data = f.read()
                    detected = chardet.detect(raw_data)
                    if detected['encoding']:
                        df = pd.read_csv(filepath, encoding=detected['encoding'])
                        print(f"[OK] CSV read successfully with detected encoding: {detected['encoding']}")
                        return df
            except:
                pass
                
        else:
            # Excel readers fallback chain
            try:
                df = pd.read_excel(filepath, engine='openpyxl')
                print("[OK] Excel read successfully with openpyxl")
                return df
            except:
                try:
                    df = pd.read_excel(filepath, engine='xlrd')
                    print("[OK] Excel read successfully with xlrd")
                    return df
                except:
                    pass
        
        raise Exception("Unable to read the file")
        
    except Exception as e:
        raise Exception(f"Error reading file: {str(e)}")


HEADER_FILL = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
HEADER_FONT = Font(color='1E293B', bold=True)
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
TEXT_ALIGNMENT = Alignment(horizontal='left', vertical='center')
NUMBER_ALIGNMENT = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='DDE1E6'),
    right=Side(style='thin', color='DDE1E6'),
    top=Side(style='thin', color='DDE1E6'),
    bottom=Side(style='thin', color='DDE1E6')
)

def _format_summary_value(value):
    if value is None or value == '':
        return '-'
    if isinstance(value, (int, float, np.integer, np.floating)):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return str(value)
        if np.isnan(number) or np.isinf(number):
            return '-'
        if abs(number - round(number)) < 1e-9:
            return f"{int(round(number)):,}"
        formatted = f"{number:,.4f}".rstrip('0').rstrip('.')
        return formatted if formatted else '0'
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return str(value)

def _auto_fit_columns(worksheet, min_width=12, max_width=60):
    for column_cells in worksheet.columns:
        column_cells = list(column_cells)
        if not column_cells:
            continue
        first_cell = column_cells[0]
        column_letter = getattr(first_cell, 'column_letter', None)
        if not column_letter:
            column_letter = get_column_letter(first_cell.column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            value = str(cell.value)
            if '\n' in value:
                length = max(len(line) for line in value.splitlines())
            else:
                length = len(value)
            if length > max_length:
                max_length = length
        adjusted_width = max(min_width, min(max_length + 4, max_width))
        worksheet.column_dimensions[column_letter].width = adjusted_width

def _style_excel_sheet(worksheet, *, freeze_pane='A2', numeric_formats=None):
    if worksheet is None:
        return
    if worksheet.max_row == 0:
        return
    if freeze_pane:
        worksheet.freeze_panes = freeze_pane
    header_rows = list(worksheet.iter_rows(min_row=1, max_row=1))
    if not header_rows:
        return
    header_row = header_rows[0]
    header_lookup = {}
    for cell in header_row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        if cell.value is not None:
            header_lookup[str(cell.value).strip().lower()] = getattr(cell, 'column', cell.column)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float, np.integer, np.floating)) and not (isinstance(cell.value, float) and (np.isnan(cell.value) or np.isinf(cell.value))):
                cell.alignment = NUMBER_ALIGNMENT
            else:
                cell.alignment = TEXT_ALIGNMENT
            cell.border = THIN_BORDER
    if numeric_formats:
        for header_name, number_format in numeric_formats.items():
            key = header_name.strip().lower()
            column_index = header_lookup.get(key)
            if not column_index:
                continue
            column_letter = get_column_letter(column_index)
            for cell in worksheet[column_letter][1:]:
                if isinstance(cell.value, (int, float, np.integer, np.floating)):
                    cell.number_format = number_format
                    cell.alignment = NUMBER_ALIGNMENT
    _auto_fit_columns(worksheet)



def create_download_files(results, filename):
    "Create download files (Excel and CSV)."
    try:
        timestamp_dt = datetime.now()
        timestamp = timestamp_dt.strftime('%Y%m%d_%H%M%S')
        base_name = filename.rsplit('.', 1)[0] if '.' in filename else filename

        results_dict = results or {}
        if not isinstance(results_dict, dict):
            results_dict = dict(results_dict) if hasattr(results, 'items') else {}

        output_files = {}

        rules_df = pd.DataFrame(results_dict.get('rulesTable') or [])
        if rules_df.empty:
            rules_df = pd.DataFrame(columns=['Antecedents', 'Consequents', 'Support', 'Confidence', 'Lift'])

        preferred_columns = ['Antecedents', 'Consequents', 'Support', 'Confidence', 'Lift']
        primary_columns = [col for col in preferred_columns if col in rules_df.columns]
        extra_columns = [col for col in rules_df.columns if col not in primary_columns]
        ordered_columns = primary_columns + extra_columns if primary_columns or extra_columns else rules_df.columns.tolist()
        export_rules_df = rules_df[ordered_columns] if ordered_columns else rules_df

        excel_filename = f"association_rules_{base_name}_{timestamp}.xlsx"
        excel_filepath = os.path.join(UPLOAD_FOLDER, excel_filename)

        with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
            export_rules_df.to_excel(writer, sheet_name='Association Rules', index=False)
            rules_ws = writer.sheets.get('Association Rules')
            _style_excel_sheet(rules_ws, freeze_pane='A2', numeric_formats={'Support': '0.000000', 'Confidence': '0.000000', 'Lift': '0.000000'})

            workbook = writer.book
            try:
                workbook.active = workbook.sheetnames.index('Association Rules')
            except Exception:
                workbook.active = 0

        output_files['excel'] = excel_filename

        csv_filename = f"association_rules_{base_name}_{timestamp}.csv"
        csv_filepath = os.path.join(UPLOAD_FOLDER, csv_filename)
        export_rules_df.to_csv(csv_filepath, index=False, encoding='utf-8-sig')
        output_files['csv'] = csv_filename

        return True, output_files

    except Exception as e:
        print(f"[ERROR] Error creating download files: {str(e)}")
        return False, str(e)





@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'message': 'Market Basket Analysis API is running successfully'
    })

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Upload an Excel/CSV file and return JSON data."""
    try:
        print("[INFO] Received upload request")
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'File type not allowed. Please upload Excel or CSV files.'}), 400
        
        print(f"[INFO] Processing file: {file.filename}")
        
        # Save uploaded file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{secure_filename(file.filename)}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        print(f"[INFO] File saved as: {filename}")
        
        # Read data from the file
        df = read_file_safely(filepath)
        
        # Basic cleanup
        df = df.fillna('')
        
        # Convert data to a list of lists for the frontend
        headers = [str(col) for col in df.columns.tolist()]
        rows = []
        
        # Limit preview output to the first 1000 rows
        display_df = df.head(1000)
        
        for _, row in display_df.iterrows():
            row_data = []
            for value in row:
                if pd.isna(value) or value is None:
                    row_data.append('')
                else:
                    row_data.append(str(value))
            rows.append(row_data)
        
        data = [headers] + rows
        
        print(f"[INFO] File processed successfully. Total rows: {len(df)}, Display rows: {len(rows)}, Columns: {len(headers)}")
        
        response_data = {
            'success': True,
            'filename': filename,
            'data': data,
            'rows': len(df),
            'columns': len(headers),
            'column_names': headers,
            'file_size': os.path.getsize(filepath)
        }
        
        return jsonify(response_data)
            
    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        print(f"[ERROR] Upload error: {str(e)}")
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500

@app.route('/api/process', methods=['POST'])
def process_data():
    """Run the market basket analysis workflow."""
    try:
        print("[INFO] Starting Market Basket Analysis...")
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        filename = data.get('filename')
        selected_columns = data.get('selectedColumns', [])
        
        print(f"[INFO] Processing: {filename}, Columns: {len(selected_columns)}")
        
        if not filename:
            return jsonify({'error': 'Missing filename'}), 400
        
        # Read the uploaded file
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        # Load the entire dataset
        df = read_file_safely(filepath)
        print(f"[INFO] File read successfully. Shape: {df.shape}")
        
        # Slice to the requested columns when provided
        if selected_columns:
            try:
                selected_df = df.iloc[:, selected_columns]
                print(f"[INFO] Selected columns: {selected_df.columns.tolist()}")
            except IndexError:
                return jsonify({'error': 'Selected columns are out of range'}), 400
        else:
            selected_df = df
        
        # Run the basket analysis
        print(f"[INFO] Running Market Basket Analysis...")
        
        analyzer = BasketAnalyzer(min_support=0.001, min_lift=1.0)
        results = analyzer.analyze_basket(selected_df, df)
        
        if not results.get('success'):
            return jsonify({'error': results.get('error', 'Analysis failed')}), 500
        
        print("[INFO] Analysis completed")
        
        # Create downloadable files
        print("[INFO] Creating download files...")
        success, output_files = create_download_files(results, filename)
        
        if not success:
            print(f"[WARN] Could not create all download files: {output_files}")
            output_files = {}
        

        if isinstance(results, dict):
            try:
                existing_output_files = results.get('outputFiles') if isinstance(results.get('outputFiles'), dict) else {}
                merged_output_files = {**existing_output_files, **output_files}
                results['outputFiles'] = merged_output_files
                results['downloadUrls'] = {**output_files}
            except Exception as merge_error:
                print(f"[WARN] Could not attach download metadata to results: {merge_error}")

        print("[INFO] Processing completed successfully")

        response_payload = {
            'success': True,
            'analysisType': 'basket',
            'selectedColumns': selected_columns,
            'results': results,
            'outputFiles': output_files,
            'downloadUrls': output_files,
            'summary': {
                'totalRows': len(df),
                'processedRows': len(selected_df),
                'selectedColumns': len(selected_columns) if selected_columns else len(df.columns),
                'totalColumns': len(df.columns),
                'processingTime': 'Completed',
                'completedAt': datetime.now().isoformat()
            }
        }

        return jsonify(response_payload)


    except Exception as e:
        logger.error(f"Processing error: {str(e)}", exc_info=True)
        return jsonify({'error': f'Processing failed: {str(e)}'}), 500



@app.route('/api/download/<format>/<path:filename>', methods=['GET'])
def download_file(format, filename):
    # Return the requested analysis file for download.
    try:
        valid_formats = {
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'csv': 'text/csv'
        }

        if format not in valid_formats:
            return jsonify({'error': 'Unsupported download format'}), 400

        decoded_filename = unquote(filename)
        safe_path = safe_join(UPLOAD_FOLDER, decoded_filename)

        if not safe_path or not os.path.exists(safe_path):
            logger.warning('Download request missing file: %s', filename)
            return jsonify({'error': 'File not found'}), 404

        download_name = os.path.basename(decoded_filename)

        return send_file(
            safe_path,
            as_attachment=True,
            download_name=download_name,
            mimetype=valid_formats[format]
        )

    except Exception as e:
        logger.error(f"Download error: {str(e)}", exc_info=True)
        return jsonify({'error': f'Download failed: {str(e)}'}), 500

# Remove outdated files
def cleanup_old_files():
    """Remove files that are older than one hour."""
    try:
        current_time = datetime.now().timestamp()
        max_age = 3600  # 1 hour
        
        for filename in os.listdir(UPLOAD_FOLDER):
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.isfile(filepath):
                file_age = current_time - os.path.getmtime(filepath)
                if file_age > max_age:
                    os.remove(filepath)
                    print(f"[INFO] Deleted old file: {filename}")
    except Exception as e:
        print(f"[WARN] Cleanup error: {str(e)}")

if __name__ == '__main__':
    print("[INFO] Starting Market Basket Analysis Server...")
    print("[INFO] Market Basket Analysis API is ready!")
    print("[INFO] Frontend can connect to: http://localhost:5000")
    print("[INFO] Upload folder:", UPLOAD_FOLDER)
    
    # Remove outdated files
    cleanup_old_files()
    
    # Start the server
    app.run(debug=True, host='0.0.0.0', port=5000)

